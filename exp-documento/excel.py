#Pega os dados do remetente e preeenche em uma planilha.

import os
import re
import pdfplumber
from openpyxl import Workbook, load_workbook

PASTA_PDFS = r"C:\PDFs_SEI_número do processo"
ARQUIVO_EXCEL = r"C:\PDFs_SEI_número do processo\dados_extraidos.xlsx"


def extrair_endereco(linha):
    linha = linha.strip()

    resultado = {
        "endereco": "",
        "numero": "",
        "complemento": "",
        "bairro": ""
    }

    linha = linha.replace("  ", " ")

    m = re.search(r"(.*?)(?:Nº|nº|No|n°)\s*(\d+)(.*?)-\s*(.+)", linha)
    if m:
        resultado["endereco"] = m.group(1).strip().rstrip(",")
        resultado["numero"] = m.group(2).strip()
        resultado["complemento"] = m.group(3).strip().strip(",")
        resultado["bairro"] = m.group(4).strip()
        return resultado

    m = re.match(r"(.+?)-(\d+)\s*-\s*(.+)", linha)
    if m:
        resultado["endereco"] = m.group(1).strip()
        resultado["numero"] = m.group(2).strip()
        resultado["bairro"] = m.group(3).strip()
        return resultado

    if "," in linha and "-" in linha:
        try:
            partes = linha.split(",", 1)
            resultado["endereco"] = partes[0].strip()

            resto = partes[1].strip()
            numero, bairro = resto.split("-", 1)

            resultado["numero"] = numero.strip()
            resultado["bairro"] = bairro.strip()

            m = re.match(r"(\d+)\s*(.*)", resultado["numero"])
            if m:
                resultado["numero"] = m.group(1)
                resultado["complemento"] = m.group(2).strip()

            return resultado
        except:
            pass

    m = re.match(r"(.+?),\s*(\d+)$", linha)
    if m:
        resultado["endereco"] = m.group(1).strip()
        resultado["numero"] = m.group(2).strip()
        return resultado

    m = re.match(r"(.+?)\s+(\d+)\s+(.+)", linha)
    if m:
        resultado["endereco"] = m.group(1).strip()
        resultado["numero"] = m.group(2).strip()
        resultado["bairro"] = m.group(3).strip()
        return resultado

    m = re.match(r"(.+?)\s+s/?n\.?", linha, flags=re.IGNORECASE)
    if m:
        resultado["endereco"] = m.group(1).strip()
        resultado["numero"] = "S/N"
        return resultado

    return resultado


def extrair_informacoes(texto):
    dados = {
        "numero_oficio": "",
        "numero_processo": "",
        "nome_destinatario": "",
        "endereco": "",
        "numero": "",
        "complemento": "",
        "bairro": "",
        "cidade": "",
        "uf": "",
        "cep": ""
    }

    linhas = [l.strip() for l in texto.split("\n") if l.strip()]

    # OFÍCIO
    for l in linhas:
        m = re.search(r"OFÍCIO SEI Nº\s*([A-Za-z0-9\/\.-]+)", l)
        if m:
            dados["numero_oficio"] = m.group(1)
            break

    # PROCESSO SEI
    for l in linhas:
        m = re.search(r"\d{5}\.\d{6}/\d{4}-\d{2}", l)
        if m:
            dados["numero_processo"] = m.group(0)
            break

    # DESTINATÁRIO COMPLETO (do "Ao" até antes do endereço)
    indice_destinatario = -1
    nome_linhas = []

    for i, l in enumerate(linhas):
        if l.startswith("Ao "):
            indice_destinatario = i
            nome_linhas.append(l)

            for j in range(i + 1, len(linhas)):
                prox = linhas[j]

                if (
                    re.search(r"\d{5}-\d{3}", prox) or
                    re.search(r"\b(Rua|Avenida|Av\.|Travessa|Praça|Alameda)\b", prox, re.IGNORECASE) or
                    re.search(r"\b\d{1,5}\b", prox)
                ):
                    break

                nome_linhas.append(prox)

            break

    dados["nome_destinatario"] = " ".join(nome_linhas)

    # ENDEREÇO
    endereco_linha = ""
    if indice_destinatario >= 0:
        for i in range(indice_destinatario + 1, min(indice_destinatario + 8, len(linhas))):
            l = linhas[i]

            if re.search(r"\d{5}-\d{3}", l):
                break

            r = extrair_endereco(l)
            if r["endereco"]:
                dados["endereco"] = r["endereco"]
                dados["numero"] = r["numero"]
                dados["complemento"] = r["complemento"]
                dados["bairro"] = r["bairro"]
                endereco_linha = l
                break

    # CEP / CIDADE / UF
    cep_index = -1
    for i, l in enumerate(linhas):
        m = re.search(r"(\d{5}-\d{3})[, ]+(.*?)(?:\s*-\s*([A-Z]{2}))?$", l)
        if m:
            dados["cep"] = m.group(1)
            dados["cidade"] = m.group(2).strip()
            if m.group(3):
                dados["uf"] = m.group(3)
            cep_index = i
            break

    if dados["uf"] == "" and cep_index >= 0 and cep_index + 1 < len(linhas):
        prox = linhas[cep_index + 1]
        if re.match(r"^[A-Z]{2}$", prox):
            dados["uf"] = prox

    if dados["bairro"] == "" and cep_index > 0:
        linha_antes = linhas[cep_index - 1]
        if "-" in linha_antes:
            possivel = linha_antes.split("-")[-1].strip()
            if len(possivel.split()) <= 4:
                dados["bairro"] = possivel

    if dados["numero"] == "":
        m = re.search(r"\b(\d{1,5})\b", endereco_linha)
        if m:
            dados["numero"] = m.group(1)

    if dados["numero"] == "":
        dados["numero"] = "S/N"

    return dados


# =================== EXCEL ===================

if not os.path.exists(ARQUIVO_EXCEL):
    wb = Workbook()
    ws = wb.active
    ws.append(["DOCUMENTO", "NOME", "ENDEREÇO", "CEP"])
    wb.save(ARQUIVO_EXCEL)

wb = load_workbook(ARQUIVO_EXCEL)
ws = wb.active

for arquivo in sorted(os.listdir(PASTA_PDFS)):
    if not arquivo.lower().endswith(".pdf"):
        continue

    caminho = os.path.join(PASTA_PDFS, arquivo)
    print(f"Extraindo: {arquivo}")

    texto = ""
    with pdfplumber.open(caminho) as pdf:
        for page in pdf.pages:
            tx = page.extract_text()
            if tx:
                texto += tx + "\n"

    dados = extrair_informacoes(texto)

    documento = f"OFÍCIO SEI Nº {dados['numero_oficio']}/{dados['numero_processo']}"

    endereco_completo = ", ".join(
        item for item in [
            dados["endereco"],
            dados["numero"] if dados["numero"] != "S/N" else None,
            dados["complemento"],
            dados["bairro"]
        ] if item
    )

    cep_completo = " - ".join(
        item for item in [
            dados["cep"],
            f"{dados['cidade']}/{dados['uf']}" if dados["cidade"] and dados["uf"] else None
        ] if item
    )

    ws.append([
        documento,
        dados["nome_destinatario"],
        endereco_completo,
        cep_completo
    ])

wb.save(ARQUIVO_EXCEL)
print("Concluído com sucesso!")
